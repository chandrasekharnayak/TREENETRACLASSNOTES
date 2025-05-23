'''
files are two type:-
text files:- txt,word,csv,excel
binary files:- image(jpg,jpeg,png),video,audio

types pf handel:-
read, write(overlapping,continue)

how to handel the file
st1:- open a file first :- open()
st2:- put the file path
st3:- put the mode
st4:- based on the requirement using the functionality

types of modes?
modes are 3 types
read ---- r
write ----- w ----overlapping
append---a ----- continue

types of read
read()
read(n)
readline()
readlines()

'''
#Escape char
# print('this is a \t boy')
# \n:- new line
# \b :- back space
# \v :- vertical tab
# \h :- horizontal tab
# \t :- tab

#how to handel a file snippet code
# '''var = open('T:\\TREENETRA\\Treenetra_class_notes\\Treenetra_AT-4\\class_2_features_of_python_and_version.txt','r')
# data = var.read()
# print(data)'''
#
# #read mode
#
# '''Read mode use for read the files data only. If when we read the file and file path is incoorect or file is not present in given location then it's
# Raise FileNotFoundError. Read mode not manupulate the files element, just for read purpose it's used.
#
# read:- read the whole files data
# read(n) :- numbers of char want to count
# readline:- read a single line at a single, from next it read next line.
# '''
#
# # var = open('T:\\TREENETRA\\Treenetra_class_notes\\Treenetra_AT-4\\class_2_features_of_python_and_version.txt','r')
# # data = var.read()
# # print(data)
#
# # data = var.read(2)
# # print(data)
#
# #readline
# '''data1 = var.readline()
# print(data1)
#
# data2 = var.readline()
# print(data2)'''
#
# #task
# """def func():
#     var = open('T:\\TREENETRA\\Treenetra_class_notes\\Treenetra_AT-4\\class_2_features_of_python_and_version.txt', 'r')
#     yield var.readline()
#
# obj = func()
#
# for i in range(10):
#     print(next(obj))"""
#
# #readlines
#
# '''var = open('T:\\TREENETRA\\Treenetra_class_notes\\Treenetra_AT-4\\class_2_features_of_python_and_version.txt', 'r')
# data = var.readlines()
#
# for i in data:
#     print(i)'''
#
#
# '''
# Write mode:- w :-
# use for write data in a file, if a file have some previous data write override the dta , previous data automatically deleted and wrire the new data.
# if the file is new then only write the new data.
# if file is not exit in the given path, wrire mode create file in same path and wrire the data
#
#  write
#  writelines
#
#  Append :- a:-
#
#  append is same as write, but it is not override the pervious data, previous data remains same, after that it's write the new data.
#  if the file is new then only write the new data.
#  if file is not exit in the given path, wrire mode create file in same path and wrire the data
#
#  write
#  writelines
#  '''
# """data = [
#     '1.use for write data in a file, if a file have some previous data write overlaping the dta , previous data automatically deleted and wrire the new data.\n',
#    '2.if the file is new then only write the new data.\n' ,
#     '3.if file is not exit in the given path, wrire mode create file in same path and wrire the data'
# ]
# var = open('T:\\FileHandelingTest\\class_19_list.txt','w')
# # var.write('use for write data in a file, if a file have some previous data write overlaping the dta , previous data automatically deleted and wrire the new data.')
# var.writelines(data)"""
#
# """var = open('T:\\FileHandelingTest\\class_19_list.txt','a')
# var.write('\nHi all, how are you, hope everything is good')
# """
#
# #how to close a file
# """
# var = open(r'T:\data_csv_handel.csv','a')
# var.close()
# print(var.closed)"""
#
# #what is with statements
# '''
# with is python keyword  and as well as it's statement. with is taken care open a file and close the file.
# help of with we will handel multiples file in a single time
#
# with open(path,mode) as variable:
#     statement
#     file open here
# automatically file code
#
# with open(path,mode) as variable1, open(path,mode) as variable2......:
#     statement
#
# '''
#
# """with open('T:\\FileHandelingTest\\class_19_list.txt','r')  as file1, open(r'T:\FileHandelingTest\class_9_10_string_basic_methods.txt','a') as file2:
#     data1 = file1.readlines()
#     file2.writelines(data1)"""


#How to handel the csv file.

import csv

import pandas as pd

# with  open('T:\\FileHandelingTest\\data_csv_write.csv','a',newline='') as file:
#     w = csv.writer(file)
#     w.writerow(['Eno','Ename','Esal','EAddr'])
#     data = int(input('Enter number of employess'))
#
#     for i in range(data):
#         eno = int(input('Enter your employee number:-'))
#         ename = input('Enter your employee name:-')
#         esal = int(input('Enter your employee salary:-'))
#         eadd = input('Enter your employee address:-')
#         w.writerow([eno,ename,esal,eadd])
#
# print('Total employees data entered')

'''with open('T:\\FileHandelingTest\\data_csv_write.csv','r') as file:
    read = csv.reader(file)
    data = list(read)
    print(data)
    for i in data:
        for j in i:
            print(j,'\t',end= '')
        print()'''

#Handel the excel file with help of pandas lib.

#excel write
import pandas
# print(dir(pandas))

'''data = {
    'Name':['Gourav','Vishal','Sumit','Lalit','Bishnu','Sourav','Kishan','Chiku','Shiva','Chiki','Pipi'],
    'Age':[23,25,27,24,25,21,27,25,27,18,16],
    'City':['Mumbai','Pune','Kolhapur','Mangalore','Udupi','Pendicherry','Bhubaneswar','Bhubaneswar','Bhubaneswar','Bhubaneswar','Bhubaneswar']
}

df = pandas.DataFrame(data)
# print(df)
df.to_excel('ouput.xlsx',index=False)'''


#excel read

import pandas

# df = pd.read_excel(r'C:\Users\admin\PycharmProjects\TreenetraClass\TREENETRA_AT_11\AdvancePython\ouput.xlsx')
# print(df)#whole data
# print(df.head())# first 5 data
# print(df.tail())# last 5 data

#handel excel with openpyxl lib

from openpyxl import load_workbook, Workbook

#excel write
'''#Create a new workbook and select for the active
wb = Workbook()
# select the active worksheet
ws = wb.active

data = [
['Name','Age','Salary','Dept'],
    ['Gourav',25,89000,'QA'],
    ['Sourav',26,91000,'Dev']
]

for i in data:
    ws.append(i)

wb.save('openpyxl.xlsx')'''

#load the workbook

'''wb = load_workbook('openpyxl.xlsx')
ws = wb.active

for i in ws.iter_rows(values_only=True):
    print(i)
'''

#Zipping and Unzipping Files

# import  zipfile
# print(dir(zipfile))
'''ZIP_DEFLATED--- create a zip file
ZIP_STORED --- zip to unzip use '''

# synatx :- ZipFile('name your providing for zip','w',ZIP_DEFLATED)
#create a zip
'''from zipfile import *

var = ZipFile('files.zip','w',ZIP_DEFLATED)
var.write('file_handeling.py')
var.write('openpyxl.xlsx')'''

#zip to unzip

from zipfile import *

file = ZipFile('files.zip','r',ZIP_STORED)
file_name = file.namelist()
for i in file_name:
    f1 = open(i,'r')
    data = f1.read()
    print(data)


















